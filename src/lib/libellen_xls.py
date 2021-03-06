from typing import List, Set, Dict, Tuple, Optional
import sys, os
import json
from datetime import datetime, time, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.drawing.image import Image
from .libellen_core import Config, Candidate, GImage, dump_b64_img_to_file, resize_image

_CONFIG: Config = None
_XLSNAME: str = "ellen.xlsx"
_WORKBOOK: openpyxl.Workbook = None
_SHEET_BAP = "People"
_SHEET_IVAR = "Entries"
_IMAGE_HEIGHT = 64

def _getXLSPath() -> str:
    if _CONFIG is None:
        return  os.path.join(".", _XLSNAME)
    else:
        return os.path.join(_CONFIG.OUTPUT_PATH, _XLSNAME)

def _open_workbook() -> bool:
    """Opens the workbook - returns true if loading was successful
     and assigned the workbook to the global _WORKBOOK.
     Will create the workbook at the XLSPATH if it does not exist."""
    global _WORKBOOK
    xlsPath = _getXLSPath()
    if _WORKBOOK is not None:
        _WORKBOOK.close() # needed in case we replace the workbook at runtime via some outside source
        _WORKBOOK = openpyxl.load_workbook(xlsPath)
        return True # already loaded, therefore true
    p = os.path.dirname(xlsPath)
    os.makedirs(p, exist_ok=True)
    try:
        _WORKBOOK = openpyxl.load_workbook(xlsPath)
        return True
    except FileNotFoundError:
        _WORKBOOK = openpyxl.Workbook()
        return _save_workbook()

def _save_workbook() -> bool:
    """ save the workbook. True is successful, False otherwise. May throw exceptions """
    if _WORKBOOK is not None:
        _WORKBOOK.save(_getXLSPath())
        return True
    return False

def _check_xls_exists() -> bool:
    """Checks that the excel file exits"""
    return os.path.isfile(_getXLSPath())

def _remove_old_xls() -> None:
    """removes any invalid or corrupt xls file we had for whatever reason. """
    try:
        print(f"Removing old db at {_getXLSPath()}")
        os.remove(_getXLSPath())
        return
    except FileNotFoundError:
        return
    except Exception as e:
        print(f"Failed to remove file at: {_getXLSPath()}: {e}")
        raise e

def _ensure_workbook() -> bool:
    """ ensures we have a valid workbook, populated with our sheets and headers. Returns true if successful, false otherwise """
    global _WORKBOOK
    if not _check_xls_exists():
        _WORKBOOK = None
        if not _open_workbook():
            return False
        # Make sure our sheets are named properly
        ws: openpyxl.worksheet.worksheet.Worksheet = _WORKBOOK.active
        ws.title = _SHEET_IVAR # first-run workbooks have an active sheet by name of Sheet1. We rename it to our IVARSheet
        ws.append(("GorillaId", "Timestamp", "Event Type", "PersonId", "Confidence", "Image", "FullBlob"))
        _WORKBOOK.create_sheet(_SHEET_BAP) # create the BAPSheet, which should be the second sheet, for usability reasons
        ws = _WORKBOOK.get_sheet_by_name(_SHEET_BAP)
        ws.append(("BAPId", "Display Name"))
        _save_workbook()
    return False

def _check_worksheet_exist(sheetname: str) -> bool:
    """given a worksheet name, check that it exists """
    wb: openpyxl.Workbook = openpyxl.load_workbook(_getXLSPath())
    return sheetname in wb

def _point_to_pixel(pt: int) -> int:
    """ converts an excel point to a pixel """
    return int(96/(72*pt))

def _pixel_to_point(px: int) -> int:
    """ converts a pixel to an excel point """
    return px * 72 / 96

def _get_image_path(img: GImage, gid: str) -> str:
    """ returns the path on disk for where to save a GImage """
    return  os.path.join(_CONFIG.SAVE_PATH, "img", f"{gid}.{img.Ext.lower()}")

def _dump_and_resize_image(img: GImage, gid: str, square_size_px: int) -> Image:
    """ takes b64 encoded image data, and saves it to disk in the square dimensions supplied. Returns an openpyxl Image """
    path = _get_image_path(img, gid)
    dump_b64_img_to_file(img.B64, path)
    resize_image(path, _IMAGE_HEIGHT)
    eimg: Image = Image(path)
    return eimg

def _delete_images_with_anchors(images: List[Image], anchors: List[str]):
    """ Given a list of anchors-to-be-removed, delete images with matching anchors """
    imgs = []
    for i,img in enumerate(images):
        real_anchor = _img_anchor_to_coordinate(img)
        if real_anchor in anchors:
            imgs.append(img)
    for img in imgs:
        images.remove(img)
    return

def _shift_images(images: List[Image], row_delta: int = 0, col_delta: int = 0):
    """ shifts all images by the amount specified """
    for img in images:
        real_anchor = _img_anchor_to_coordinate(img)
        coord = openpyxl.utils.coordinate_to_tuple(real_anchor)
        anchor = (coord[0] + row_delta, coord[1] + col_delta)
        new_coord = f"{openpyxl.utils.get_column_letter(anchor[1])}{anchor[0]}"
        img.anchor = new_coord
    return

def _img_anchor_to_coordinate(img: Image) -> str:
    """ takes the image anchor and returns its excel coordinate, like F3 """
    if isinstance(img.anchor, str):
        return img.anchor # sometimes an anchor may already be in coordinator form
    ianch = img.anchor._from
    anc_col = openpyxl.utils.get_column_letter(ianch.col+1) # internal field is 0-indexed
    real_anchor = f"{anc_col}{ianch.row+1}" # internal field is 0-indexed
    return real_anchor

def _clean_image(img: GImage, gid: str):
    """ removes the downloaded GImage. Its only use was to be added to the excel sheet """
    if img is not None:
        path = _get_image_path(img, gid)
        try:
            os.remove(path)
        except:
            return
    return


def update_bap(candidates: List[Candidate]):
    """ given a list of potential candidate matches, update the bap sheet to ensure that any new known people are properly inserted"""
    if not candidates:
        return
    _open_workbook()
    sheet = _WORKBOOK.get_sheet_by_name(_SHEET_BAP)
    for c in candidates:
        should_add = True
        for row in sheet.iter_rows():
            if row[0].value == c.Id:
                should_add = False
                break
        if should_add:
            sheet.append((c.Id, c.DisplayName))
    _save_workbook()
    return

def update_ivar(gorillaId: str, timestamp: datetime, eventType: str, img: GImage, candidate: Candidate, jobj: str):
    """ Inserts data into the Ivar entries sheet. """
    _open_workbook()
    sheet = _WORKBOOK.get_sheet_by_name(_SHEET_IVAR)
    gid = gorillaId.replace("{", "").replace("}", "")
    pid = candidate.Id if candidate else None
    score = candidate.SimiliarityScore if candidate else None
    eimg: Image = None

    sheet.append((gorillaId, timestamp, eventType, pid, score, None, jobj)) #image slot is None because we need to special insert it in the next step
    rc = sheet.max_row
    if img:
        eimg = _dump_and_resize_image(img, gid, _IMAGE_HEIGHT)
        sheet.add_image(eimg, f"F{rc}")
    sheet.row_dimensions[rc].height = _pixel_to_point(_IMAGE_HEIGHT) # set all row heights to be the image height

    _save_workbook()
    #_clean_image(img, gid) # remove any saved images, because once their written to the worksheet, they're stored inside the xls file
    return

def ensure() -> bool:
    return _ensure_workbook()

def set_config(config: Config):
    global _CONFIG
    _CONFIG = config
    return

def prune_old_data() -> str:
    """ removes old data according to the Config Rules. 
        If any pruning needs to happen, the file gets rolled over into one with the format of
        Ellen_YYYY_mm_DD.count.xlsx 
        
        returns the name of the rolled over file if exists, or None if it didnt rollover"""
    _ensure_workbook()
    _open_workbook()
    sheet = _WORKBOOK.get_sheet_by_name(_SHEET_IVAR)
    rowcount = sheet.max_row
    rollover = False

    # check that the file has at least 2 rows. One for header, one for first data
    if rowcount < 2:
        return None
    
    # check if MAX_RECORDS has been exceeded
    excess_rows = rowcount - _CONFIG.MAX_RECORD_COUNT -1 # -1 because we want to be mindful of the header row
    if excess_rows > 0:
        print(f"Current file has exceeded the configured rowcount. Max: {_CONFIG.MAX_RECORD_COUNT}, Got: {rowcount}")
        rollover = True
    
    # check if first data row has exceeded MAX_KEEP_DAYS
    max_date = datetime.now() - timedelta(days=_CONFIG.MAX_KEEP_DAYS)
    earliest_entry = sheet[2][1].value
    if earliest_entry <= max_date:
        print(f"Current file had records which exceed the configured Keep Day limit: Max Keep Days: {_CONFIG.MAX_KEEP_DAYS}, earliest record: {(datetime.now() - earliest_entry).days} ago ({earliest_entry.strftime('%Y-%m-%d')})")
        rollover = True

    # check if MAX_SIZE has been exceeded
    max_bytes = int(_CONFIG.MAX_SIZE * 1e6)
    fsize = os.path.getsize(_getXLSPath())
    if fsize > max_bytes:
        print(f"Current file has exceeded the configured MAX_SIZE: Configured size: {_CONFIG.MAX_SIZE}, file size: {fsize}")
        rollover = True

    # perform the rename and rollover
    # NF TODO
    if rollover:
        name = _rollover()
        print(f"The prevous ellen.xlsx file has been rolled over. It is available under the filename {name}")
        return name

    return None

def _rollover() -> str:
    """ moves the current ellen.xlsx to a rolled over 'Ellen-YYYY-dd-MM.c.xlsx' file.
    Returns the name of the rolled over file
    """
    _WORKBOOK.close() # ensure that the file lock is released
    today = datetime.now()
    new_fname_first = f"ellen-{today.strftime('%Y-%m-%d')}"
    fpath = Path(_getXLSPath())
    num_others = len([x for x in os.listdir(_CONFIG.OUTPUT_PATH) if x.startswith(new_fname_first)])+1
    new_fname = f"{new_fname_first}.{num_others}.xlsx"
    fpath.rename(os.path.join(os.path.dirname(fpath), new_fname))
    return new_fname

def main():
    print("lib_elen_XLS:")
    print(f"Checking {_getXLSPath()} exists: {_check_xls_exists()}")
    print("Ensuring XLS creation...")
    _ensure_workbook()
    print("Pruning old data")
    pruned = prune_old_data()
    print(f"Pruned {pruned} old rows")
    return 0

if __name__ == "__main__":
    sys.exit(main())